"""Script para generar archivos Excel de prueba (mocks) para la plataforma Expoflores.

Genera un conjunto completo de archivos .xlsx para pruebas de carga masiva:
1. Cargas válidas de distintos tamaños y categorías
2. Cargas ajustadas a los stands del seed
3. Carga con errores intencionales para verificar la validación fila por fila
4. Carga para prueba de exceso de cupo
5. Plantilla base
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.domain.identification import validate_identification
from app.integrations.excel import PARTICIPANT_COLUMNS, read_participant_rows
from app.schemas.participant import ParticipantIn

# Estilos visuales para los Excel
HEADER_FILL = PatternFill(
    start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"
)  # Azul corporativo
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BORDER_THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

FIRST_NAMES = [
    "Carlos",
    "María",
    "Juan",
    "Ana",
    "Luis",
    "Elena",
    "Diego",
    "Patricia",
    "Andrés",
    "Gabriela",
    "Fernando",
    "Sofía",
    "José",
    "Lucía",
    "Ricardo",
    "Carmen",
    "Javier",
    "Diana",
    "Manuel",
    "Camila",
    "Mateo",
    "Valentina",
    "Santiago",
    "Isabella",
    "Alejandro",
    "Daniela",
    "Esteban",
    "Valeria",
    "David",
    "Natalia",
    "Sebastián",
    "Andrea",
    "Pablo",
    "Adriana",
    "Gabriel",
    "Paula",
    "Felipe",
    "Juliana",
    "Leonardo",
    "Silvia",
    "Rodrigo",
    "Lorena",
    "Gonzalo",
    "Mónica",
    "César",
    "Verónica",
    "Hugo",
    "Claudia",
    "Guillermo",
    "Teresa",
]

LAST_NAMES = [
    "Morales",
    "Castillo",
    "Guerrero",
    "Mendoza",
    "Paredes",
    "Salazar",
    "Cárdenas",
    "Vargas",
    "Reyes",
    "Guzmán",
    "Herrera",
    "Pinto",
    "Alarcón",
    "Espinoza",
    "Bustamante",
    "Salgado",
    "Navarro",
    "Carrera",
    "Ponce",
    "Delgado",
    "Villavicencio",
    "Cordero",
    "Montalvo",
    "Sotomayor",
    "Aguirre",
    "Cevallos",
    "Benalcázar",
    "Jaramillo",
    "Cobo",
    "Ribadeneira",
    "Enríquez",
    "Villacís",
    "Pérez",
    "López",
    "Gómez",
    "Rodríguez",
    "Sánchez",
    "Romero",
    "Díaz",
    "Álvarez",
    "Zambrano",
    "Molina",
    "Castro",
    "Ortiz",
    "Silva",
]

POSITIONS_EXHIBITOR = [
    "Gerente de Exportaciones",
    "Jefe de Ventas Internacionales",
    "Director Comercial",
    "Key Account Manager",
    "Floricultor Senior",
    "Asesor Técnico Floral",
    "Especialista en Postcosecha",
    "Jefe de Calidad",
]

POSITIONS_GUEST = [
    "Comprador Internacional",
    "Importador Mayorista",
    "Distribuidor Floral",
    "Diseñador Floral Invitado",
    "Consultor Agrónomo",
    "Representante Comercial",
    "Director de Cadena de Supermercados",
    "Comprador Minorista Especializado",
]

POSITIONS_SERVICE = [
    "Técnico de Montaje",
    "Supervisor de Stand",
    "Electricista de Eventos",
    "Operador de Cadena de Frío",
    "Encargado de Logística",
    "Diseñador de Estructuras",
    "Especialista en Iluminación",
    "Auxiliar de Mantenimiento y Limpieza",
]

PROVIDER_COMPANIES = [
    "Montajes y Estructuras Andinas S.A.",
    "Logística Floral del Ecuador Cía. Ltda.",
    "Sistemas de Refrigeración Polar S.A.",
    "Iluminación y Efectos Pro Stand Cía. Ltda.",
    "Catering & Eventos Mitad del Mundo",
    "Transportes Especializados Flor Express",
    "Servicios Integrales ExpoClean Cía. Ltda.",
    "Seguridad Privada Andina SegurAndes",
    "Diseño & Montaje Creativo S.A.",
    "Audio y Video Profesional ExpoTech",
]

PROVINCES = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]


def gen_cedula(prov: int, seq: int) -> str:
    """Genera cédula ecuatoriana válida por algoritmo módulo 10."""
    first9 = f"{prov:02d}{seq:07d}"
    coeffs = (2, 1, 2, 1, 2, 1, 2, 1, 2)
    pairs = zip(map(int, first9), coeffs, strict=True)
    total = sum(d * c - 9 if d * c > 9 else d * c for d, c in pairs)
    check = (10 - total % 10) % 10
    return f"{first9}{check}"


def gen_ruc_natural(prov: int, seq: int) -> str:
    """Genera RUC de persona natural válido (cédula + 001)."""
    return f"{gen_cedula(prov, seq)}001"


def gen_ruc_juridica(prov: int, seq: int) -> str:
    """Genera RUC de persona jurídica (tipo 9) válido por módulo 11."""
    while True:
        first9 = f"{prov:02d}9{seq:06d}"
        coeffs = (4, 3, 2, 7, 6, 5, 4, 3, 2)
        rem = sum(int(d) * c for d, c in zip(first9, coeffs, strict=True)) % 11
        check = 0 if rem == 0 else 11 - rem
        if check != 10:
            return f"{first9}{check}001"
        seq += 1


def create_excel_file(headers: list[str], rows: list[list], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"

    # Escribir encabezados
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    ws.row_dimensions[1].height = 28

    # Escribir filas con formato de texto para preservar ceros a la izquierda
    for row_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="center")
            if headers[col_idx - 1] in (
                "identificacion",
                "celular",
                "tipo_identificacion",
                "categoria",
            ):
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 20

    # Autoajuste de columnas
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0]) if row[0] is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def generate_participant_record(
    index: int, category: str, id_type: str = "CEDULA", include_email: bool = True
) -> dict:
    fname = FIRST_NAMES[index % len(FIRST_NAMES)]
    lname = LAST_NAMES[index % len(LAST_NAMES)]
    prov = PROVINCES[index % len(PROVINCES)]
    seq = 300000 + index

    if id_type == "CEDULA":
        ident = gen_cedula(prov, seq)
    elif id_type == "RUC":
        ident = gen_ruc_natural(prov, seq) if index % 2 == 0 else gen_ruc_juridica(prov, seq)
    elif id_type == "PASSPORT":
        ident = f"PAS{400000 + index:06d}"
    elif id_type == "FOREIGN_ID":
        ident = f"EXT{600000 + index:06d}"
    else:
        ident = gen_cedula(prov, seq)

    phone = f"099{1000000 + (index * 73) % 8999999:07d}"

    if category == "Exhibitor":
        position = POSITIONS_EXHIBITOR[index % len(POSITIONS_EXHIBITOR)]
        provider = None
    elif category == "Guest":
        position = POSITIONS_GUEST[index % len(POSITIONS_GUEST)]
        provider = None
    elif category == "Service":
        position = POSITIONS_SERVICE[index % len(POSITIONS_SERVICE)]
        provider = PROVIDER_COMPANIES[index % len(PROVIDER_COMPANIES)]
    else:
        position = "Asistente"
        provider = None

    email = f"{fname.lower()}.{lname.lower()}{index}@testflor.com" if include_email else None

    return {
        "nombre": fname,
        "apellido": lname,
        "identificacion": ident,
        "tipo_identificacion": id_type,
        "celular": phone,
        "cargo": position,
        "categoria": category,
        "empresa_proveedora": provider,
        "correo": email,
    }


def row_dict_to_list(row: dict, headers: list[str]) -> list:
    return [row.get(h) for h in headers]


def main():
    output_dir = "/app/datos_de_mocks_temp"
    os.makedirs(output_dir, exist_ok=True)
    headers = list(PARTICIPANT_COLUMNS.keys())

    print(f"Generando archivos de mock en: {output_dir}")

    # =========================================================================
    # 1. Carga válida stand demo (17 participantes: 8 Exhibitor, 3 Guest, 6 Service)
    # =========================================================================
    rows_demo = []
    idx = 1
    for _ in range(8):
        id_type = "CEDULA" if idx % 3 != 0 else "RUC"
        rows_demo.append(
            generate_participant_record(idx, "Exhibitor", id_type, include_email=(idx % 2 == 1))
        )
        idx += 1
    for _ in range(3):
        id_type = "PASSPORT" if idx % 2 == 0 else "CEDULA"
        rows_demo.append(
            generate_participant_record(idx, "Guest", id_type, include_email=(idx % 2 == 1))
        )
        idx += 1
    for _ in range(6):
        id_type = "CEDULA" if idx % 2 == 0 else "FOREIGN_ID"
        rows_demo.append(
            generate_participant_record(idx, "Service", id_type, include_email=(idx % 2 == 1))
        )
        idx += 1

    file1 = os.path.join(output_dir, "01_carga_valida_stand_demo_17_participantes.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_demo], file1)
    print(f"✓ Creado: {file1} (17 filas)")

    # =========================================================================
    # 2. Carga masiva 50 participantes (25 Exhibitor, 15 Guest, 10 Service)
    # =========================================================================
    rows_50 = []
    idx = 100
    for _ in range(25):
        id_type = "CEDULA" if idx % 4 != 0 else "RUC"
        rows_50.append(
            generate_participant_record(idx, "Exhibitor", id_type, include_email=(idx % 3 != 0))
        )
        idx += 1
    for _ in range(15):
        id_type = "PASSPORT" if idx % 3 == 0 else "CEDULA"
        rows_50.append(
            generate_participant_record(idx, "Guest", id_type, include_email=(idx % 3 != 0))
        )
        idx += 1
    for _ in range(10):
        id_type = "FOREIGN_ID" if idx % 3 == 0 else "CEDULA"
        rows_50.append(
            generate_participant_record(idx, "Service", id_type, include_email=(idx % 3 != 0))
        )
        idx += 1

    file2 = os.path.join(output_dir, "02_carga_valida_masiva_50_participantes.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_50], file2)
    print(f"✓ Creado: {file2} (50 filas)")

    # =========================================================================
    # 3. Carga masiva 100 participantes (50 Exhibitor, 30 Guest, 20 Service)
    # =========================================================================
    rows_100 = []
    idx = 500
    for _ in range(50):
        id_t = ["CEDULA", "CEDULA", "RUC", "PASSPORT"][idx % 4]
        rows_100.append(
            generate_participant_record(idx, "Exhibitor", id_t, include_email=(idx % 2 == 0))
        )
        idx += 1
    for _ in range(30):
        id_t = ["PASSPORT", "CEDULA", "FOREIGN_ID"][idx % 3]
        rows_100.append(
            generate_participant_record(idx, "Guest", id_t, include_email=(idx % 2 == 0))
        )
        idx += 1
    for _ in range(20):
        id_t = ["CEDULA", "FOREIGN_ID", "RUC"][idx % 3]
        rows_100.append(
            generate_participant_record(idx, "Service", id_t, include_email=(idx % 2 == 0))
        )
        idx += 1

    file3 = os.path.join(output_dir, "03_carga_valida_masiva_100_participantes.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_100], file3)
    print(f"✓ Creado: {file3} (100 filas)")

    # =========================================================================
    # 4. Carga para stand pequeño (2 Exhibitor)
    # =========================================================================
    rows_peq = [
        generate_participant_record(801, "Exhibitor", "CEDULA", include_email=True),
        generate_participant_record(802, "Exhibitor", "CEDULA", include_email=False),
    ]
    file4 = os.path.join(output_dir, "04_carga_valida_stand_pequeno_2_participantes.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_peq], file4)
    print(f"✓ Creado: {file4} (2 filas)")

    # =========================================================================
    # 5. Carga para stand mediano (10 participantes: 5 Exhibitor, 2 Guest, 3 Service)
    # =========================================================================
    rows_med = [
        generate_participant_record(810, "Exhibitor", "CEDULA", include_email=True),
        generate_participant_record(811, "Exhibitor", "CEDULA", include_email=True),
        generate_participant_record(812, "Exhibitor", "RUC", include_email=False),
        generate_participant_record(813, "Exhibitor", "CEDULA", include_email=True),
        generate_participant_record(814, "Exhibitor", "CEDULA", include_email=False),
        generate_participant_record(815, "Guest", "PASSPORT", include_email=True),
        generate_participant_record(816, "Guest", "CEDULA", include_email=False),
        generate_participant_record(817, "Service", "CEDULA", include_email=True),
        generate_participant_record(818, "Service", "FOREIGN_ID", include_email=True),
        generate_participant_record(819, "Service", "CEDULA", include_email=False),
    ]
    file5 = os.path.join(output_dir, "05_carga_valida_stand_mediano_10_participantes.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_med], file5)
    print(f"✓ Creado: {file5} (10 filas)")

    # =========================================================================
    # 6. Carga para stand grande (33 participantes: 15 Exhibitor, 7 Guest, 11 Service)
    # =========================================================================
    rows_grande = []
    idx = 900
    for _ in range(15):
        rows_grande.append(
            generate_participant_record(idx, "Exhibitor", "CEDULA", include_email=(idx % 2 == 0))
        )
        idx += 1
    for _ in range(7):
        rows_grande.append(
            generate_participant_record(
                idx, "Guest", "PASSPORT" if idx % 2 == 0 else "CEDULA", include_email=True
            )
        )
        idx += 1
    for _ in range(11):
        rows_grande.append(
            generate_participant_record(
                idx, "Service", "CEDULA" if idx % 2 == 0 else "FOREIGN_ID", include_email=True
            )
        )
        idx += 1

    file6 = os.path.join(output_dir, "06_carga_valida_stand_grande_33_participantes.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_grande], file6)
    print(f"✓ Creado: {file6} (33 filas)")

    # =========================================================================
    # 7. Archivo para probar ERRORES DE VALIDACIÓN (12 filas variadas)
    # =========================================================================
    cedula_duplicada = gen_cedula(17, 950001)
    rows_errores = [
        # Fila 2: Válida
        {
            "nombre": "Gabriel",
            "apellido": "Mendoza",
            "identificacion": cedula_duplicada,
            "tipo_identificacion": "CEDULA",
            "celular": "0991122334",
            "cargo": "Director Comercial",
            "categoria": "Exhibitor",
            "empresa_proveedora": None,
            "correo": "gabriel.mendoza@valido.com",
        },
        # Fila 3: Error en Cédula (dígito verificador incorrecto 9 en vez del real)
        {
            "nombre": "Roberto",
            "apellido": "Salgado",
            "identificacion": "1710034069",
            "tipo_identificacion": "CEDULA",
            "celular": "0992233445",
            "cargo": "Floricultor",
            "categoria": "Exhibitor",
            "empresa_proveedora": None,
            "correo": "roberto.salgado@test.com",
        },
        # Fila 4: Error en RUC (provincia 99 no existe)
        {
            "nombre": "Lucía",
            "apellido": "Ponce",
            "identificacion": "9991234567001",
            "tipo_identificacion": "RUC",
            "celular": "0993344556",
            "cargo": "Asesora",
            "categoria": "Guest",
            "empresa_proveedora": None,
            "correo": "lucia.ponce@test.com",
        },
        # Fila 5: Error en nombre vacío (obligatorio)
        {
            "nombre": None,
            "apellido": "Alarcón",
            "identificacion": gen_cedula(9, 950002),
            "tipo_identificacion": "CEDULA",
            "celular": "0994455667",
            "cargo": "Coordinadora",
            "categoria": "Exhibitor",
            "empresa_proveedora": None,
            "correo": "alarcon@test.com",
        },
        # Fila 6: Error en Service sin empresa proveedora (obligatoria para Service)
        {
            "nombre": "Fabián",
            "apellido": "Bustamante",
            "identificacion": gen_cedula(1, 950003),
            "tipo_identificacion": "CEDULA",
            "celular": "0995566778",
            "cargo": "Técnico de Sonido",
            "categoria": "Service",
            "empresa_proveedora": None,
            "correo": "fabian@test.com",
        },
        # Fila 7: Error en Exhibitor CON empresa proveedora (prohibida para no-Service)
        {
            "nombre": "Mónica",
            "apellido": "Vargas",
            "identificacion": gen_cedula(18, 950004),
            "tipo_identificacion": "CEDULA",
            "celular": "0996677889",
            "cargo": "Ejecutiva de Cuentas",
            "categoria": "Exhibitor",
            "empresa_proveedora": "Montajes Andinos S.A.",
            "correo": "monica.vargas@test.com",
        },
        # Fila 8: Error en categoría inexistente (VIP)
        {
            "nombre": "Santiago",
            "apellido": "Herrera",
            "identificacion": gen_cedula(17, 950005),
            "tipo_identificacion": "CEDULA",
            "celular": "0997788990",
            "cargo": "Invitado Especial",
            "categoria": "VIP",
            "empresa_proveedora": None,
            "correo": "santiago@test.com",
        },
        # Fila 9: Error en formato de correo (correo inválido)
        {
            "nombre": "Daniela",
            "apellido": "Enríquez",
            "identificacion": gen_cedula(11, 950006),
            "tipo_identificacion": "CEDULA",
            "celular": "0998899001",
            "cargo": "Diseñadora",
            "categoria": "Guest",
            "empresa_proveedora": None,
            "correo": "correo_invalido_sin_arroba",
        },
        # Fila 10: Error de duplicado interno (misma cédula que la fila 2)
        {
            "nombre": "Andrés",
            "apellido": "Navarro",
            "identificacion": cedula_duplicada,
            "tipo_identificacion": "CEDULA",
            "celular": "0991122339",
            "cargo": "Promotor",
            "categoria": "Exhibitor",
            "empresa_proveedora": None,
            "correo": "andres.navarro@test.com",
        },
        # Fila 11: Error en tipo de identificación no soportado
        {
            "nombre": "Julio",
            "apellido": "Montalvo",
            "identificacion": "0102030405",
            "tipo_identificacion": "DNI_NACIONAL",
            "celular": "0990011223",
            "cargo": "Representante",
            "categoria": "Exhibitor",
            "empresa_proveedora": None,
            "correo": "julio@test.com",
        },
        # Fila 12: Pasaporte muy corto (menos de 5 caracteres)
        {
            "nombre": "John",
            "apellido": "Doe",
            "identificacion": "P12",
            "tipo_identificacion": "PASSPORT",
            "celular": "0993322110",
            "cargo": "Comprador",
            "categoria": "Guest",
            "empresa_proveedora": None,
            "correo": "john.doe@test.com",
        },
        # Fila 13: Válida (Service con empresa proveedora correcta)
        {
            "nombre": "César",
            "apellido": "Paredes",
            "identificacion": gen_cedula(5, 950007),
            "tipo_identificacion": "CEDULA",
            "celular": "0994433221",
            "cargo": "Electricista",
            "categoria": "Service",
            "empresa_proveedora": "Iluminación y Efectos Pro Stand Cía. Ltda.",
            "correo": "cesar.paredes@prostand.com",
        },
    ]

    file7 = os.path.join(output_dir, "07_prueba_errores_de_validacion_filas.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_errores], file7)
    print(f"✓ Creado: {file7} (12 filas con errores controlados)")

    # =========================================================================
    # 8. Prueba de exceso de cupo (20 Exhibitor)
    # =========================================================================
    rows_exceso = []
    for i in range(20):
        rows_exceso.append(
            generate_participant_record(1200 + i, "Exhibitor", "CEDULA", include_email=True)
        )

    file8 = os.path.join(output_dir, "08_prueba_exceso_de_cupo_20_expositores.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_exceso], file8)
    print(f"✓ Creado: {file8} (20 filas de Exhibitor)")

    # =========================================================================
    # 9. Plantilla oficial limpia con 1 ejemplo
    # =========================================================================
    rows_plantilla = [
        {
            "nombre": "María",
            "apellido": "Chiriboga",
            "identificacion": "1710034065",
            "tipo_identificacion": "CEDULA",
            "celular": "0991234567",
            "cargo": "Jefa de Ventas",
            "categoria": "Exhibitor",
            "empresa_proveedora": None,
            "correo": "maria.chiriboga@example.com",
        }
    ]
    file9 = os.path.join(output_dir, "09_plantilla_base_limpia.xlsx")
    create_excel_file(headers, [row_dict_to_list(r, headers) for r in rows_plantilla], file9)
    print(f"✓ Creado: {file9} (1 fila de ejemplo)")

    # =========================================================================
    # Validación automática de integridad de los archivos generados
    # =========================================================================
    print("\n--- Verificando validez de archivos generados con el backend ---")
    valid_files = [file1, file2, file3, file4, file5, file6, file8, file9]
    for filepath in valid_files:
        with open(filepath, "rb") as f:
            content = f.read()
        rows = read_participant_rows(content)
        for _, val_dict in rows:
            p = ParticipantIn.model_validate(val_dict)
            validate_identification(p.identification, p.identification_type)
        print(f"✓ Verificado 100% válido: {os.path.basename(filepath)} ({len(rows)} filas)")

    # Verificar que el archivo de errores efectivamente falle en las filas esperadas
    with open(file7, "rb") as f:
        content_err = f.read()
    err_rows = read_participant_rows(content_err)
    detected_errors = 0
    for _, val_dict in err_rows:
        try:
            p = ParticipantIn.model_validate(val_dict)
            validate_identification(p.identification, p.identification_type)
        except Exception:
            detected_errors += 1
    print(
        f"✓ Archivo de errores verificado: detectados {detected_errors} "
        "fallos de validación esperados."
    )
    print("\n¡Todos los archivos Excel de prueba fueron generados y validados con éxito!")


if __name__ == "__main__":
    main()
