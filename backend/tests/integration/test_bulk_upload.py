"""Carga masiva de credenciales desde Excel (F11, punto extra E2a).

Trazabilidad: R14. Identificaciones ficticias, validas por algoritmo.
"""

from io import BytesIO
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.domain.identification import _modulo10_digit
from app.integrations.excel import PARTICIPANT_COLUMNS
from app.models import Participant, User
from tests.conftest import auth_headers

BULK = "/api/v1/me/participants/bulk"
TEMPLATE = "/api/v1/me/participants/template.xlsx"


def cedula(serial: int) -> str:
    """Cedula ficticia pero valida por algoritmo: provincia 17 y digito verificador real."""
    base = f"17{serial:07d}"
    return base + str(_modulo10_digit(base))


def row(identification: str, **overrides: Any) -> dict[str, Any]:
    data: dict[str, Any] = {
        "nombre": "Ana",
        "apellido": "Torres",
        "identificacion": identification,
        "tipo_identificacion": "CEDULA",
        "celular": "0990000001",
        "cargo": "Personal de stand",
        "categoria": "Exhibitor",
        "empresa_proveedora": None,
        "correo": None,
    }
    data.update(overrides)
    return data


def workbook_bytes(rows: list[dict[str, Any]], headers: list[str] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    columns = headers if headers is not None else list(PARTICIPANT_COLUMNS)
    sheet.append(columns)
    for item in rows:
        sheet.append([item.get(column) for column in columns])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def upload(
    client: TestClient,
    rep: User,
    content: bytes,
    dry_run: bool = False,
    filename: str = "credenciales.xlsx",
) -> Any:
    return client.post(
        BULK,
        params={"dry_run": str(dry_run).lower()},
        files={"file": (filename, content, "application/octet-stream")},
        headers=auth_headers(rep),
    )


def count(db: Session) -> int:
    return int(db.execute(select(func.count()).select_from(Participant)).scalar_one())


# --- R14: filas invalidas, cero inserciones ---------------------------------------------------


def test_invalid_rows_report_and_no_inserts(client: TestClient, db: Session, rep_a: User) -> None:
    """40 filas, 3 invalidas: se reporta cada una por numero de fila y no entra ninguna."""
    rows = [row(cedula(i)) for i in range(40)]
    rows[4]["identificacion"] = "1710000010"  # digito verificador incorrecto
    rows[7]["nombre"] = None  # obligatorio
    rows[9]["categoria"] = "Invitado"  # fuera del enum

    response = upload(client, rep_a, workbook_bytes(rows))

    assert response.status_code == 422
    body = response.json()
    assert body["code"] == "BULK_UPLOAD_INVALID_ROWS"
    details = body["details"]
    assert details["total_rows"] == 40
    # La cabecera es la fila 1: la fila i del lote es la i+2 del Excel.
    assert {e["row"] for e in details["errors"]} == {6, 9, 11}
    assert details["valid_rows"] == 37
    assert count(db) == 0


def test_duplicate_within_file_detected(client: TestClient, db: Session, rep_a: User) -> None:
    """: los duplicados DENTRO del archivo tambien se reportan, no solo los de la base."""
    rows = [row(cedula(0)), row(cedula(1)), row(cedula(0), nombre="Otra")]

    response = upload(client, rep_a, workbook_bytes(rows))

    assert response.status_code == 422
    errors = response.json()["details"]["errors"]
    assert len(errors) == 1
    assert errors[0]["row"] == 4
    assert "fila 2" in errors[0]["message"]
    assert count(db) == 0


def test_duplicate_against_database_reports_owner(
    client: TestClient, db: Session, rep_a: User, rep_b: User
) -> None:
    assert upload(client, rep_b, workbook_bytes([row(cedula(0))])).status_code == 200

    response = upload(client, rep_a, workbook_bytes([row(cedula(0))]))

    assert response.status_code == 422
    assert "Flores del Valle" in response.json()["details"]["errors"][0]["message"]
    assert count(db) == 1


def test_dry_run_inserts_nothing(client: TestClient, db: Session, rep_a: User) -> None:
    """: mismo informe con y sin `dry_run`; la unica diferencia es el COMMIT."""
    content = workbook_bytes([row(cedula(0)), row(cedula(1))])

    preview = upload(client, rep_a, content, dry_run=True).json()
    assert preview == {"total_rows": 2, "valid_rows": 2, "inserted": 0, "dry_run": True}
    assert count(db) == 0

    confirmed = upload(client, rep_a, content).json()
    assert confirmed == {"total_rows": 2, "valid_rows": 2, "inserted": 2, "dry_run": False}
    assert count(db) == 2


def test_quota_is_checked_against_the_whole_batch(
    client: TestClient, db: Session, rep_a: User
) -> None:
    """25 m2 dan 4 credenciales Guest. Cinco filas caben una a una, pero el lote no."""
    rows = [row(cedula(i), categoria="Guest") for i in range(5)]

    response = upload(client, rep_a, workbook_bytes(rows))

    assert response.status_code == 409
    body = response.json()
    assert body["code"] == "QUOTA_EXCEEDED"
    assert body["details"] == {"category": "Guest", "quota": 4, "used": 0, "requested": 5}
    assert count(db) == 0


def test_service_row_requires_provider_company(client: TestClient, rep_a: User) -> None:
    rows = [row(cedula(0), categoria="Service")]

    errors = upload(client, rep_a, workbook_bytes(rows)).json()["details"]["errors"]

    assert errors[0]["row"] == 2
    assert errors[0]["field"] == "empresa_proveedora"


def test_optional_email_is_imported(client: TestClient, db: Session, rep_a: User) -> None:
    """: la columna `correo` es opcional; si viene, se guarda."""
    rows = [row(cedula(0), correo="ana@example.com"), row(cedula(1))]

    assert upload(client, rep_a, workbook_bytes(rows)).status_code == 200

    emails = set(db.execute(select(Participant.email)).scalars())
    assert emails == {"ana@example.com", None}


# --- Rechazo del archivo antes de parsearlo --------------------------------------------


def test_csv_renamed_to_xlsx_is_rejected(client: TestClient, rep_a: User) -> None:
    content = b"nombre,apellido\nAna,Torres\n"

    response = upload(client, rep_a, content, filename="credenciales.xlsx")

    assert response.status_code == 422
    assert response.json()["code"] == "VALIDATION_ERROR"


def test_wrong_extension_is_rejected(client: TestClient, rep_a: User) -> None:
    content = workbook_bytes([row(cedula(0))])
    assert upload(client, rep_a, content, filename="credenciales.xls").status_code == 422


def test_oversized_file_is_rejected(client: TestClient, rep_a: User) -> None:
    response = upload(client, rep_a, b"PK\x03\x04" + b"0" * (2 * 1024 * 1024))
    assert response.status_code == 422
    assert "MB" in response.json()["message"]


def test_missing_columns_are_reported(client: TestClient, rep_a: User) -> None:
    content = workbook_bytes([row(cedula(0))], headers=["nombre", "apellido"])

    response = upload(client, rep_a, content)

    assert response.status_code == 422
    assert "identificacion" in response.json()["message"]


def test_blank_rows_are_ignored(client: TestClient, db: Session, rep_a: User) -> None:
    content = workbook_bytes([row(cedula(0)), {}, row(cedula(1))])

    assert upload(client, rep_a, content).json()["total_rows"] == 2
    assert count(db) == 2


def test_numeric_cells_are_read_as_text(client: TestClient, db: Session, rep_a: User) -> None:
    """Excel entrega 1710000017 como numero; la identificacion no puede llegar como '1.7e9'."""
    content = workbook_bytes([row(int(cedula(0)), celular=990000001)])

    assert upload(client, rep_a, content).status_code == 200

    participant = db.execute(select(Participant)).scalar_one()
    assert participant.identification == cedula(0)
    assert participant.phone == "990000001"


# --- Plantilla ---------------------------------------------------------------------------


def test_template_has_the_exact_columns(client: TestClient, rep_a: User) -> None:
    response = client.get(TEMPLATE, headers=auth_headers(rep_a))

    assert response.status_code == 200
    sheet = load_workbook(BytesIO(response.content)).active
    headers = [cell.value for cell in sheet[1]]
    assert headers == list(PARTICIPANT_COLUMNS)


def test_bulk_upload_requires_representative(client: TestClient, admin_user: User) -> None:
    response = upload(client, admin_user, workbook_bytes([row(cedula(0))]))
    assert response.status_code == 403
