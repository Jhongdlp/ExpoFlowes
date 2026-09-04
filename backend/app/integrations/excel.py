"""Generacion de Excel con openpyxl.

Las columnas por categoria salen de las reglas del evento, no de una lista fija: si manana
una feria define una cuarta categoria de credencial, el reporte la incluye sin tocar codigo.
"""

from collections.abc import Sequence
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_BASE_HEADERS = [
    "Razon social",
    "Identificacion tributaria",
    "Nombre del stand",
    "Metraje (m2)",
    "Categoria de stand",
]


def _autosize(sheet: Any, headers: Sequence[str]) -> None:
    for index, header in enumerate(headers, start=1):
        width = max(len(header) + 2, 12)
        sheet.column_dimensions[get_column_letter(index)].width = min(width, 40)


def exhibitors_report(
    summaries: Sequence[dict[str, Any]], categories: Sequence[str], event_name: str
) -> bytes:
    """Un stand por fila, con cuota / asignadas / disponibles de cada categoria."""
    headers = list(_BASE_HEADERS)
    for category in categories:
        headers += [f"{category}: cuota", f"{category}: asignadas", f"{category}: disponibles"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expositores"
    sheet["A1"] = event_name
    sheet["A1"].font = Font(bold=True, size=13)
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for item in summaries:
        row: list[Any] = [
            item["legal_name"],
            item["tax_id"],
            item["stand_name"],
            item["requested_m2"],
            item["stand_category"],
        ]
        for category in categories:
            row += [
                item["quota"].get(category, 0),
                item["assigned"].get(category, 0),
                item["available"].get(category, 0),
            ]
        sheet.append(row)

    sheet.freeze_panes = "A4"
    _autosize(sheet, headers)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- Carga masiva de participantes (F11) --------------------------------------------------

# Encabezado exacto del Excel -> campo del esquema. El orden del dict ES el orden de las
# columnas de la plantilla (§13): un solo sitio define ambos, no pueden divergir.
PARTICIPANT_COLUMNS: dict[str, str] = {
    "nombre": "first_name",
    "apellido": "last_name",
    "identificacion": "identification",
    "tipo_identificacion": "identification_type",
    "celular": "phone",
    "cargo": "position",
    "categoria": "category",
    "empresa_proveedora": "provider_company",
    "correo": "email",
}

MAX_UPLOAD_BYTES = 2 * 1024 * 1024
MAX_UPLOAD_ROWS = 500
HEADER_ROW = 1
_XLSX_MAGIC = b"PK\x03\x04"  # un .xlsx es un ZIP; un .csv renombrado no lo es


class InvalidWorkbookError(Exception):
    """El archivo no es utilizable. El servicio la traduce al error HTTP (§9.4)."""


def ensure_xlsx(filename: str | None, content: bytes) -> None:
    """Tipo y tamaño ANTES de parsear (§8.10): un archivo hostil no llega a openpyxl."""
    if not (filename or "").lower().endswith(".xlsx"):
        raise InvalidWorkbookError("El archivo debe tener extension .xlsx.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise InvalidWorkbookError(
            f"El archivo supera el maximo de {MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
        )
    if not content.startswith(_XLSX_MAGIC):
        raise InvalidWorkbookError("El archivo no es un libro de Excel .xlsx valido.")


def _cell_text(value: Any) -> str | None:
    """Excel entrega numeros donde el negocio espera texto: 1710034065 llega como int y
    0990000000 puede llegar como float. Todo se normaliza a texto sin decimal colgante."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def read_participant_rows(content: bytes) -> list[tuple[int, dict[str, str | None]]]:
    """(numero de fila del Excel, valores) por cada fila con datos.

    No valida negocio: solo convierte el archivo en filas. Quien decide si una fila es valida
    es el mismo esquema Pydantic del alta manual (§8.4).
    """
    from openpyxl import load_workbook  # import local: solo lo paga quien sube un archivo

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lanza de todo ante un archivo corrupto
        raise InvalidWorkbookError("El archivo no se pudo leer como libro de Excel.") from exc

    sheet = workbook.active
    if sheet is None:
        raise InvalidWorkbookError("El libro no tiene ninguna hoja.")

    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise InvalidWorkbookError("El archivo esta vacio.")

    positions = {
        str(name).strip().lower(): index for index, name in enumerate(header) if name is not None
    }
    missing = [h for h in PARTICIPANT_COLUMNS if h not in positions]
    if missing:
        raise InvalidWorkbookError(
            "Faltan columnas en el archivo: " + ", ".join(missing) + ". Use la plantilla."
        )

    parsed: list[tuple[int, dict[str, str | None]]] = []
    for offset, row in enumerate(rows, start=HEADER_ROW + 1):
        values = {
            field: _cell_text(
                row[positions[header_name]] if positions[header_name] < len(row) else None
            )
            for header_name, field in PARTICIPANT_COLUMNS.items()
        }
        if all(v is None for v in values.values()):
            continue  # fila en blanco: el usuario borro el contenido, no la fila
        if len(parsed) >= MAX_UPLOAD_ROWS:
            raise InvalidWorkbookError(
                f"El archivo supera el maximo de {MAX_UPLOAD_ROWS} filas por carga."
            )
        parsed.append((offset, values))

    workbook.close()
    return parsed


def participants_template() -> bytes:
    """Plantilla con las 9 columnas de §13, en orden, mas una fila de ejemplo."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participantes"
    headers = list(PARTICIPANT_COLUMNS)
    sheet.append(headers)
    for cell in sheet[HEADER_ROW]:
        cell.font = Font(bold=True)
    sheet.append(
        [
            "Maria",
            "Chiriboga",
            "1710034065",
            "CEDULA",
            "0991234567",
            "Jefa de ventas",
            "Exhibitor",
            None,
            "maria.chiriboga@example.com",
        ]
    )
    sheet.freeze_panes = "A2"
    _autosize(sheet, headers)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
